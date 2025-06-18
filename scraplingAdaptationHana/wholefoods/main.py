from scrapling.fetchers import StealthyFetcher
from urllib.parse import urlparse
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
import asyncio
import time
import httpx, json
import base64, urllib.parse
import pandas as pd
import random
import logging, sys, pathlib
import re
import unicodedata



BASE_DIR = pathlib.Path(__file__).resolve().parent
LOG_FILE = BASE_DIR / "wholefoods_logging/wholefoods_scrape.log"
SCRAPE_DUMP_FILE = BASE_DIR / "wholefoods_scrape_dump" / "catalogue_dump.json"
SCRAPE_DUMP_FILE.parent.mkdir(parents=True, exist_ok=True)
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
FINAL_RESULTS_DIR = BASE_DIR / "wholefoods_price_compare"
FINAL_RESULTS_DIR.mkdir(parents=True, exist_ok=True)


# PRICE_SHEET = input("📁 Enter path to your input XLSX file: ").strip().strip('"\'')

THRESHOLD = 85
DELTA_PCT = 0.25
DF_COMP_FI = None
cookies_for_api = {}
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logging.getLogger("httpx").setLevel(logging.WARNING)
log = logging.getLogger(__name__)

def prompt_for_excel(prompt_text: str) -> str:
    while True:
        path = input(prompt_text).strip().strip('"').strip("'")
        if path.lower().endswith(".xlsx") and pathlib.Path(path).exists():
            return path
        print("❌ Invalid file. Please drag and drop a valid .xlsx file or paste the full path.")

PRICE_SHEET2 = prompt_for_excel("Drag and drop the Whole Foods price sheet (xlsx)")

async def set_zip_wholefoods(page, zipcode):
    global cookies_for_api
    await page.click("text='Find a Store' >> visible=true")

    # html = await page.content()
    # with open("sidebar_dom.html", "w", encoding="utf-8") as f:
    #     f.write(html)
    
    #debugging 
    # frames = page.frames
    # print(f"Found {len(frames)} frames")
    # for f in frames:
    #     print("Frame name:", f.name, "URL:", f.url)



    await page.wait_for_selector("iframe[title='stores-modal']", timeout=10000)

    frame_element = await page.query_selector("iframe[title='stores-modal']")
    frame = await frame_element.content_frame()

    #frame = page.frame(name="store-web-page")
    if frame is None:
        print(" Could not find iframe 'store-web-page'") 
        return


    await frame.wait_for_selector("#store-finder-search-bar", timeout=10000)
    await frame.fill("#store-finder-search-bar", zipcode)
    await frame.wait_for_timeout(500)
    await frame.press("#store-finder-search-bar", "Enter")
    await frame.wait_for_selector(".w-store-finder-store-selector", timeout=5000)

    await frame.click("span.w-makethismystore[tabindex='0']")
    await page.wait_for_timeout(1000)

    await page.wait_for_selector("button.modalCloseButtonStyle", timeout=5000)
    await page.click("button.modalCloseButtonStyle")
    print("✅ Clicked close button inside iframe")

    target_address = "1440 P Street NW, 20005"
    found = await wait_for_address(page, target_address, timeout=5000)
    if found: 
        print(f" Address is : {target_address}")
    else:
        raise RuntimeError(f"Expected store address to contain '20005', but doesn't")
    
    cookies_for_api = await export_cookies_for_httpx(page)
    raw_cookie = cookies_for_api["wfm_store_d8"]
    raw_cookie = urllib.parse.unquote(raw_cookie)
    missing = (-len(raw_cookie) % 4)
    raw_cookie  += "=" * missing
    payload = json.loads(base64.b64decode(urllib.parse.unquote(raw_cookie)))
    print("🍪 Cookies for API:", cookies_for_api)
    
    print("✅ store id in cookie:", payload)

    await page.wait_for_timeout(3000)  # Wait for the page to update



async def main(urlstr):
    url = urlstr
    domain = urlparse(url).netloc.replace("www.", "")
    wf_zipcode = "20005"

    async def page_action(page):
        if domain in ZIP_HANDLERS:
            await ZIP_HANDLERS[domain](page, wf_zipcode)
        else:
            print(f"No handler for domain: {domain}")
        return page
    
    response = await StealthyFetcher.async_fetch(
        url=url,
        headless=False,
        network_idle=True,
        disable_resources=True,
        page_action=page_action
    )
    


async def wait_for_address(page, target_text, timeout=7000, poll_interval=500):
    start_time = time.time()
    while (time.time() - start_time) * 1000 < timeout:
        html = await page.content()
        if target_text in html:
            return True
        await page.wait_for_timeout(poll_interval)
    return False


async def fetch_wholefoods_api(cookies, category="all-products", limit=60) -> list[dict]:
    store_id = store_id_from_cookie(cookies)
    if not cookies_for_api:
        raise RuntimeError("cookies_for_api is still empty")

    store_id = "10135"
    category = "all-products"
    limit = 60
    offset = 0
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/plain, */*",
        "Referer": f"https://www.wholefoodsmarket.com/products/all-products",
        "Origin": "https://www.wholefoodsmarket.com",
        "x-requested-with": "XMLHttpRequest",
    }

    products = []
    seen_slugs = set()
    no_new_items_count = 0
    max_no_new_items = 3  # Exit after 3 consecutive API calls with no new items
    consecutive_empty_batches = 0
    first_run = True
    with httpx.Client(headers=headers, cookies=cookies_for_api, timeout=20) as c:
        while True:
            url = f"https://www.wholefoodsmarket.com/api/products/category/{category}"
            params = {
                "leafCategory": category,
                "store": store_id,
                "limit": limit,
                "offset": offset,
            }
            r = c.get(url, params=params)
            r.raise_for_status()
            data = r.json()
            raw_items = data.get("results", [])
            
            if offset == 0:
                total = total_skus_from_facets(data.get("facets", []))
                print(f"Total SKU's according to API: {total}")
                log.info(f"API reports {total} total SKUs")
            elif offset == 7000:
                total = total_skus_from_facets(data.get("facets", []))
                print(f"Total SKU's according to API: {total}")
                log.info(f"API reports {total} total SKUs")
            elif offset >= 9998:
                #offset = 0
                first_run = False
            
            if not raw_items:
                log.warning(f"Empty response at offset {offset} – stopping crawl")
                break
                
            batch = []
            for itm in raw_items:
                slug = itm.get("slug")
                if slug in seen_slugs:
                    continue
                seen_slugs.add(slug)
                item = extract_item_fields(itm)
                batch.append(item)
            
            if len(batch) == 0:
                consecutive_empty_batches += 1
                log.warning(f"No new items at offset {offset} (batch {consecutive_empty_batches})")

                if consecutive_empty_batches >= 30:
                    if limit > 48:
                        limit -= 1
                

                # Exit after several consecutive empty batches
                if consecutive_empty_batches >= max_no_new_items:
                    log.warning(f"Exiting after {consecutive_empty_batches} consecutive empty batches")
                    break
            else:
                # Reset the counter when we get new items
                consecutive_empty_batches = 0
                
            products.extend(batch)
            log.info(f"Fetched {len(batch):>3} new items (offset={offset}, cumulative={len(products)})")
            
            # Exit if we've reached or exceeded the expected total
            if len(products) >= total:
                log.info(f"Reached expected total of {total} products")
                break
                
            # Exit if we're getting too many items (safety check)
            if len(products) > total * 1.5:
                log.warning(f"Collected {len(products)} items which exceeds expected total {total} by 50%. Stopping crawl.")
                break
                
            # Still trying to figure out the exact workings of the API-- does it increase by a set amount or index by batch
            if offset + limit > 9999:
                offset = 9999
            elif offset == 0 and not first_run:
                offset = offset + 1
            else:
                offset += limit
            
    log.info(f"Total unique skus: {len(products)}")
    
    save_catalogue_dump(products)

    return products


async def export_cookies_for_httpx(page, cookie_names = None):
    """
    Return a {name:value} dict of cookies for httpx, limited to `cookie_names`
    if provided.
    """
    jar = {}
    for c in await page.context.cookies():
        if not c["domain"].endswith("wholefoodsmarket.com"):
            continue
        print("🍪", c["name"])
        if (cookie_names is None) or (c["name"] in cookie_names):
            jar[c["name"]] = c["value"]
    return jar

def extract_item_fields(item: dict) -> dict:
    """
    Keep only the interesting fields, renaming them a bit on the way.
    Fields that don't exist on a given item are returned as None.
    """
    return {
        "name"            : item.get("name"),
        "brand"           : item.get("brand"),
        "regular_price"   : item.get("regularPrice"),
         ## NOTE: To activate debug mode, uncomment the comments
        # optional sale fields
        # "sale_price"          : item.get("salePrice"),
        # "incremental_sale_price": item.get("incrementalSalePrice"),
        # "sale_start_date"     : item.get("saleStartDate"),
        # "sale_end_date"       : item.get("saleEndDate"),
        "slug"          : item.get("slug"),
        # "uom"           : item.get("uom"),  
        # handy if you need the PDP URL later
        # "slug"            : item.get("slug"),
        # "store"           : item.get("store"),
        # "uom"             : item.get("uom"),
        # "is_local"        : item.get("isLocal"),
        # "image_thumb"     : item.get("imageThumbnail"),
    }

async def main_fetch_and_save(items):
    #df = pd.DataFrame(items)
    out_file = "wholefoods_products.xlsx"
    items.to_excel(out_file, index=False)
    print(f"saved to {out_file}")

def store_id_from_cookie(jar: dict) -> str:
    """Return the storeId encoded inside wfm_store_d8."""
    raw = urllib.parse.unquote(jar["wfm_store_d8"])
    raw += "=" * (-len(raw) % 4)                           # pad
    payload = json.loads(base64.urlsafe_b64decode(raw))
    return payload["id"]                                   

def total_skus_from_facets(facets:list) -> int | None:
    for facet in facets:
        if facet.get("slug") == "category":
            for ref in facet.get("refinements", []):
                if ref.get("slug") == "all-products":
                    return ref.get("count")
    return None

def save_catalogue_dump(products: list[dict]):
    """Save scraped catalogue to dump file with timestamp"""
    dump_data = {
        "scraped_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_products": len(products),
        "products": products
    }
    
    try:
        with open(SCRAPE_DUMP_FILE, 'w', encoding='utf-8') as f:
            json.dump(dump_data, f, indent=2, ensure_ascii=False)
        log.info(f"📁 Catalogue dump saved to {SCRAPE_DUMP_FILE}")
        print(f"📁 Catalogue dump saved: {len(products)} products → {SCRAPE_DUMP_FILE}")
    except Exception as e:
        log.error(f"Failed to save catalogue dump: {e}")

def load_catalogue_dump() -> list[dict] | None:
    """Load previously scraped catalogue from dump file"""
    if not SCRAPE_DUMP_FILE.exists():
        return None
    
    try:
        with open(SCRAPE_DUMP_FILE, 'r', encoding='utf-8') as f:
            dump_data = json.load(f)
        
        scraped_at = dump_data.get("scraped_at", "unknown")
        products = dump_data.get("products", [])
        
        log.info(f"📂 Loaded catalogue dump: {len(products)} products (scraped: {scraped_at})")
        print(f"📂 Loaded catalogue dump: {len(products)} products (scraped: {scraped_at})")
        
        return products
    except Exception as e:
        log.error(f"Failed to load catalogue dump: {e}")
        return None


def normalise(txt: str | None) -> str:
    """lower-case, strip, collapse whitespace."""
    if not txt:
        return ""
    txt = str(txt)
    return " ".join(txt.lower().split())

def build_matcher(df_cat: pd.DataFrame, brand_included: bool) -> callable:
    """Return helper that fuzzy-matches a query to df_comp rows."""
    ### TODO: if brand not included, then separate the brand from the slug and separate the brand from the Item Description, 
    ### Then do another similarity matcher across the brands. There is a slight functional error right now where non-branded 
    ### items are the highest scoring items due to extraneous words being removed. Separating the brands will help with this. 
    if not brand_included:
        choices = df_cat["base"].tolist()           # list[str]
        idx_map = {v: i for i, v in enumerate(choices)}

        def best_match(query: str):
            """Return (row_idx_in_comp, score)."""
            match, score, _ = process.extractOne(
                query,
                choices,
                scorer=fuzz.token_set_ratio,         # order-insensitive!
            )
            return idx_map[match], score
        return best_match
    else:
        sizes = df_cat["size_norm"].tolist()
        names = df_cat["name"].tolist()
        brands = df_cat["brand"].tolist()

        by_brand: dict[str, list[str]] = {}
        by_size: dict[str, list[int]] = {}
        for idx, s in enumerate(sizes):
            by_size.setdefault(s, []).append(idx)
        for idx, b in enumerate(brands):
            by_brand.setdefault(b, []).append(idx)

        def match(q_name: str, q_size: str, b_name: str):
            cand_rows = by_size.get(q_size, [])
            brand_rows = by_brand.get(b_name, [])
            ### TODO: split the conditional branch to do if not brand rows, calculate brand score
            if not cand_rows or not brand_rows:
                return None, 0.0        # no candidate with same size

            choices = [names[i] for i in cand_rows]

            # scorer 1: token_sort_ratio
            idx1, s1, _ = process.extractOne(
                q_name, choices, scorer=fuzz.token_sort_ratio)

            # scorer 2: partial_ratio
            idx2, s2, _ = process.extractOne(
                q_name, choices, scorer=fuzz.partial_ratio)
            
            idx1 = choices.index(idx1)
            idx2 = choices.index(idx2)

            if s1 >= s2:
                return cand_rows[idx1], s1
            else:
                return cand_rows[idx2], s2
        return match
    

async def compare_prices(scraped_items: list[dict]):
    # 1. load reference file
    #ToDo: get the correct names from the two excel files
    df_src = pd.read_excel(PRICE_SHEET2)
    df_src.columns = [col.lower() for col in df_src.columns]
    brand_included = False          # expects cols: Name, Price
    if not ("brand" in df_src.columns):
        df_src["norm"] = df_src["item desc."].map(normalise)
        df_src["size"] = df_src["size"].map(normalize_size_string)

        # 2. turn scraped list -> DataFrame
        df_comp = pd.DataFrame(scraped_items)
        #df_comp["norm"] = df_comp["slug"].map(normalise)
        ####### NOTE: split_slug() no longer needs to return base_cmp
        base_cmp, size_cmp = zip(*df_comp["slug"].map(split_slug))
        #print(f"base_cmp: {len(base_cmp)}, df_cmp: {len(df_comp)}")  
        #df_comp["base"] = list(base_cmp)
        df_comp["base"] = (
        df_comp["slug"]
        .str.replace(r"-b[0-9a-z]{9,}$", "", regex=True)  # drop trailing id
        .map(normalise)                                  # lower-case, etc.
        )
        print(df_comp["base"].head(10))
        df_comp["size"] = [s.replace(" ", "") if s else None for s in size_cmp]

        # 3. build matcher on scraped catalogue
        matcher = build_matcher(df_comp, brand_included)
    elif "brand" in df_src.columns:
        brand_included = True
        df_src["norm"] = df_src["item desc."].map(normalise)
        df_src["size_norm"] = df_src["size"].map(normalize_size_string)
        df_src["brand"] = df_src["brand"].map(normalize_brand)
        df_comp = pd.DataFrame(scraped_items)
        _, size_cmp = zip(*df_comp["slug"].map(split_slug))
        df_comp["name"] = df_comp["name"].map(normalise)
        df_comp["size"] = [s.replace(" ", "") if s else None for s in size_cmp]
        df_comp["size_norm"] = df_comp["size"].map(normalize_size_string)
        df_comp["brand"] = df_comp["brand"].map(normalize_brand)
        DF_COMP_FI = df_comp
        matcher = build_matcher(df_comp, brand_included)
    # 4. match each row from price-sheet
    best_rows   = []
    best_scores = []
    price_diffs = []
    match_info = []
    THRESHOLD = 30
    comp_price = None
    if not ("brand" in df_src.columns):
        for q_norm, size_unit, src_price in zip(df_src["norm"], df_src["size"], df_src["price"]):
            match_name = None
            comp_price = None
            potentiate_name = None
            row_idx, score = matcher(q_norm)
            best_rows.append(row_idx)
            best_scores.append(score)

            if row_idx is not None: 
                potentiate_name = df_comp.loc[row_idx, "name"]

            if row_idx is not None and (score >= THRESHOLD):
                size_src  = (size_unit or "").strip().lower()
                size_comp = (df_comp.loc[row_idx, "size"] or "").strip().lower()

                if size_src and size_src == size_comp:
                    comp_price = df_comp.loc[row_idx, "regular_price"]
                    match_name = df_comp.loc[row_idx, "name"]
                else:
                    # helpful debug:
                    #print(f"Some results-- Score>=Threshold: {score}, {THRESHOLD}, {score>=THRESHOLD}, size unit not empty: {size_unit}, \n misc: {size_unit.lower()}, {df_comp.loc[row_idx, 'size']}")
                    log.debug(
                        "size mismatch - src=%s / comp=%s  (row %s, score %.1f)",
                        size_src, size_comp, row_idx, score,
                    )
                    row_idx = None          # reject
                    score   = 0.0
            # if score >= THRESHOLD and size_unit and size_unit.lower() == df_comp.loc[row_idx, "size"].lower():
            #     comp_price = df_comp.loc[row_idx, "regular_price"]
            # else:                                     # no reliable hit
            #     print(f"Some results-- Score>=Threshold: {score}, {THRESHOLD}, {score>=THRESHOLD}, size unit not empty: {size_unit}, \n misc: {size_unit.lower()}, {df_comp.loc[row_idx, 'size'].lower()}")
            #     row_idx = None
            #     comp_price = None
            #     score = 0
            #print(df_comp.loc[row_idx, 'size'])
            # best_rows.append(row_idx)
            # #match_info.append(match_name)
            # best_scores.append(score)

            if pd.notna(src_price) and pd.notna(comp_price):
                diff_pct = abs(comp_price - src_price) / src_price
            else:
                diff_pct = None
            

            price_diffs.append({
                "MatchName" : df_comp.loc[row_idx, "name"] if row_idx is not None else "n/a",
                "CompPrice": comp_price,
                "PotetentiateName": potentiate_name,
                #"Δ pct"    : diff_pct,
            })
    elif ("brand" in df_src.columns):
        num_potential_matches = 0
        for q_norm, brand, size_unit, src_price in zip(df_src["norm"], df_src["brand"], df_src["size_norm"], df_src["price"]):
            match_name = None
            comp_price = None
            potentiate_name = None
            reject_reason = []
            row_idx, score = matcher(q_norm, size_unit, brand)
            best_rows.append(row_idx)
            best_scores.append(score)
            if row_idx is not None: 
                potentiate_name = df_comp.loc[row_idx, "name"]

            if row_idx is not None and (score >= THRESHOLD):
                size_src  = (size_unit or "").strip().lower()
                size_comp = (df_comp.loc[row_idx, "size_norm"] or "").strip().lower()
                brand_src = brand.strip()
                brand_comp = df_comp.loc[row_idx, "brand" or ""].strip().lower()

                if brand_src and brand_src == brand_comp:
                    num_potential_matches += 1
                    print(f"size_src: {size_src}, size_comp: {size_comp}\
                          \nsrcName vs compName: {q_norm}, {df_comp.loc[row_idx, 'slug']}\
                          \n num_potential_matches: {num_potential_matches}, {brand_src}, {brand_comp}")
                if size_src and size_src == size_comp and brand_src and brand_src == brand_comp:
                    comp_price = df_comp.loc[row_idx, "regular_price"]
                    match_name = df_comp.loc[row_idx, "name"]
                else:
                    #helpful debug:
                    print(f"Some results-- \
                          \nsrcName vs compName: {q_norm}, {df_comp.loc[row_idx, 'slug']}, \
                           \nScore>=Threshold: {score}, {THRESHOLD}, {score>=THRESHOLD},\
                            \nsize_comp: {size_unit.lower()}, {df_comp.loc[row_idx, 'size']},\
                             \nbrand_src: {brand_src}, \
                              \nbrand_comp: {brand_comp}")
                    log.debug(
                        "size mismatch - src=%s / comp=%s  (row %s, score %.1f)",
                        size_src, size_comp, row_idx, score,
                    )
                    if size_src != size_comp:
                        reject_reason.append(f"size mismatch - src={size_src} / comp={size_comp}")
                    if brand_src != brand_comp:
                        reject_reason.append(f"brand mismatch - src={brand_src} / comp={brand_comp}")
                    row_idx = None          # reject
                    score   = 0.0

            # if score >= THRESHOLD and size_unit and size_unit.lower() == df_comp.loc[row_idx, "size"].lower():
            #     comp_price = df_comp.loc[row_idx, "regular_price"]
            # else:                                     # no reliable hit
            #     print(f"Some results-- Score>=Threshold: {score}, {THRESHOLD}, {score>=THRESHOLD}, size unit not empty: {size_unit}, \n misc: {size_unit.lower()}, {df_comp.loc[row_idx, 'size'].lower()}")
            #     row_idx = None
            #     comp_price = None
            #     score = 0
            #print(df_comp.loc[row_idx, 'size'])
            # best_rows.append(row_idx)
            # #match_info.append(match_name)
            # best_scores.append(score)

            # if pd.notna(src_price) and pd.notna(comp_price):
            #     diff_pct = abs(comp_price - src_price) / src_price
            # else:
            #     diff_pct = None
            

            price_diffs.append({
                ## NOTE: To activate debug mode, uncomment the comments
                "Brand" : df_comp.loc[row_idx, "brand"] if row_idx is not None else "n/a",
                #"MatchName" : df_comp.loc[row_idx, "name"] if row_idx is not None else "n/a",
                #"Size": df_comp.loc[row_idx, "size"] if row_idx is not None else "n/a",
                "CompPrice": comp_price,
                #"PotetentiateName": potentiate_name,
                "RejectReason": reject_reason,
                #"Δ pct"    : diff_pct,
            })

    price_cols = pd.DataFrame(price_diffs)
    keep_cols = ["upc", 'item desc.', 'price']
    df_out = pd.concat(
        [df_src[keep_cols].reset_index(drop=True),
         price_cols,
         pd.Series(best_scores, name="Score")],
        axis=1,
    )

    # 5. flag big deltas
    # df_out["Flag"] = df_out["Δ pct"].apply(
    #     lambda x: "BIG" if (pd.notna(x) and x > DELTA_PCT) else ""
    # )

    # 6. write result
    out_file = FINAL_RESULTS_DIR / "wholefoods_pc.xlsx"
    preformating_xl_helper(df_out, out_file)
    print(f"📝 comparison report ➜ {out_file}")
    return df_comp



def split_slug(slug: str) -> tuple[str, str | None]:
    tokens = slug.split("-")
    if tokens and _AMZ_ID_RE.match(tokens[-1]):
        tokens.pop()
    
    size_val, size_unit, idx = None, None, -1
    for i, t in enumerate(tokens):
        m = _NUM_UNIT_RE.fullmatch(t)
        if m:
            size_val, size_unit, idx = m.group(1), m.group(2), i
            break
    if size_val is None:
        for i in range(len(tokens) - 1):
            if tokens[i].replace(".", "", 1).isdigit() and tokens[i + 1] in SIZE_UNITS:
                size_val, size_unit, idx = tokens[i], tokens[i + 1], i
                break

    if size_val:
        # remove size tokens from list
        del tokens[idx: idx + 2]  # safe even if second idx is OOB
        size = f"{size_val} {size_unit}".lower()
    else:
        size = None
    
    base = " ".join(tokens).replace("  ", " ").strip().lower()
    #log.info({base}, {size})
    return base, size

def split_name(txt: str) -> tuple[str, str | None]:
    """
    crude free-text → (base, size) using the same regex.
    """
    if not isinstance(txt, str):
        return "", None
    m = _NUM_UNIT_RE.search(txt)
    if m:
        size = f"{m.group(1)} {m.group(2)}".lower()
        base = _NUM_UNIT_RE.sub("", txt).strip().lower()
        return base, size
    return txt.lower(), None

def normalize_size_string(size_str: str) -> str:
    """
    Normalize size strings by extracting only numeric characters.
    - Removes units like 'oz', 'ml', 'ct'
    - Removes decimal points
    - Keeps digits only
    """
    if not isinstance(size_str, str):
        return ""

    # Remove everything that's not a digit
    digits_only = re.sub(r"\D", "", size_str)
    return digits_only


def normalize_brand(b: str) -> str:
    if not isinstance(b, str):
        return ""
    b = unicodedata.normalize("NFKD", b)  # normalize unicode
    b = b.lower()
    b = re.sub(r"[’'®™\-\.]", "", b)     
    b = re.sub(r"\s+", " ", b).strip()   
    return b


def preformating_xl_helper(df_out, out_file):
    df_out.to_excel(out_file, index=False)
    wb = load_workbook(out_file)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # 1️⃣ Find the MatchName and Score column indexes
    col_names = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
    match_name_col_idx = col_names.get("MatchName")  # Changed from "Brand" to "MatchName"
    score_col_idx = col_names.get("Score")

    if match_name_col_idx is None or score_col_idx is None:
        # Debug: print available columns to help diagnose issues
        print("Available columns:", list(col_names.keys()))
        raise ValueError("MatchName or Score column not found")
    
    match_col_letter = get_column_letter(match_name_col_idx)
    score_col_letter = get_column_letter(score_col_idx)
    
    ws.auto_filter.ref = f"{match_col_letter}1:{match_col_letter}{max_row}"
    dxf = DifferentialStyle(
        fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    )
    rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    # Note: apply to *whole row*, so highlight A to last col if score > 40
    rule.formula = [f"${score_col_letter}2>45"]
    last_col_letter = get_column_letter(max_col)
    ws.conditional_formatting.add(f"A2:{last_col_letter}{max_row}", rule)

    thick = Side(border_style="medium", color="000000")
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=match_name_col_idx, max_col=max_col):
        for cell in row:
            cell.border = Border(
                left=thick if cell.column == match_name_col_idx else None,
                right=thick if cell.column == max_col else None,
                top=thick if cell.row == 1 else None,
                bottom=thick if cell.row == max_row else None,
            )

    wb.save(out_file)
    wb.close()




ZIP_HANDLERS = {
    "wholefoodsmarket.com": set_zip_wholefoods,
}

# Potential size units to look for in the product name
SIZE_UNITS = {
    "oz", "fl", "floz", "fl-oz", "g", "mg", "kg",
    "ml", "l", "lb", "lbs", "ct", "pack", "pk", "fz",
}

# Regex patterns for matching ASIN and size units
_AMZ_ID_RE   = re.compile(r"^[a-z0-9]{10}$", re.I)
_NUM_UNIT_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*-?\s*(" + "|".join(SIZE_UNITS) + r")\b", re.I)

if __name__ == "__main__":
    import sys
    
    # Check for command line arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == "use-dump":
            # Use existing catalogue dump instead of scraping
            items = load_catalogue_dump()
            if items is None:
                print("❌ No catalogue dump found. Run without arguments to scrape first.")
                sys.exit(1)
            print("🔄 Using existing catalogue dump for price comparison...")
            DF_COMP_FI = asyncio.run(compare_prices(items))
            asyncio.run(main_fetch_and_save(DF_COMP_FI))
            sys.exit(0)
        else:
            print("Usage:")
            print("  python main.py                - Full scrape and compare")
            print("  python main.py use-dump       - Use existing dump for comparison")
            sys.exit(1)
    
    # Default behavior: full scrape
    url = "https://www.wholefoodsmarket.com"
    asyncio.run(main(url))
    items = asyncio.run(fetch_wholefoods_api(cookies_for_api))
    
    DF_COMP_FI = asyncio.run(compare_prices(items))
    asyncio.run(main_fetch_and_save(DF_COMP_FI))